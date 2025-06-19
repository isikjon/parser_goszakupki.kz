#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Экспортер данных goszakup в красивый XLSX формат
"""

import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class XLSXExporter:
    def __init__(self, db_file: str = 'turbo_goszakup.db'):
        self.db_file = db_file
        
    def export_demo_to_xlsx(self, csv_file: str = 'demo_suppliers.csv', output_file: str = 'goszakup_suppliers_full.xlsx'):
        """Экспорт демо данных из CSV в красивый XLSX"""
        logger.info(f"Экспорт демо данных из {csv_file} в {output_file}")
        
        # Читаем CSV
        df = pd.read_csv(csv_file, encoding='utf-8-sig')
        
        # Создаем workbook
        wb = Workbook()
        
        # Основной лист с данными
        ws_main = wb.active
        ws_main.title = "Поставщики"
        
        # Определяем структуру колонок
        main_columns = {
            'participant_number': 'Номер участника',
            'name': 'Наименование',
            'bin': 'БИН',
            'iin': 'ИИН', 
            'rnn': 'РНН',
            'supplier_id': 'ID поставщика'
        }
        
        # Основная информация
        basic_columns = {
            'detail_Дата регистрации': 'Дата регистрации',
            'detail_Дата последнего обновления': 'Дата обновления',
            'detail_Роли участника': 'Роли участника',
            'detail_Состоит в реестре государственных заказчиков': 'В реестре гос. заказчиков',
            'detail_Наименование на рус. языке': 'Наименование (рус)',
            'detail_Наименование на каз. языке': 'Наименование (каз)',
            'detail_Резиденство': 'Резиденство',
            'detail_КАТО': 'КАТО',
            'detail_Регион': 'Регион'
        }
        
        # Контактная информация
        contact_columns = {
            'detail_E-Mail:': 'Email',
            'detail_Контактный телефон:': 'Телефон',
            'detail_Вебсайт:': 'Веб-сайт'
        }
        
        # Документы
        doc_columns = {
            'detail_Серия свидетельства (для ИП) и номер свидетельства о государственной регистрации': 'Серия и номер свидетельства',
            'detail_Дата свидетельства о государственной регистрации': 'Дата свидетельства',
            'detail_Наименование администратора(ов) отчетности': 'Администратор отчетности'
        }
        
        # Подготавливаем данные для основного листа
        main_data = []
        all_columns = {**main_columns, **basic_columns, **contact_columns, **doc_columns}
        
        for _, row in df.iterrows():
            row_data = []
            for col_key, col_name in all_columns.items():
                value = row.get(col_key, '')
                if pd.isna(value):
                    value = ''
                row_data.append(str(value))
            main_data.append(row_data)
        
        # Заголовки
        headers = list(all_columns.values())
        
        # Записываем данные
        ws_main.append(headers)
        for row_data in main_data:
            ws_main.append(row_data)
        
        # Форматирование основного листа
        self._format_main_sheet(ws_main, len(headers))
        
        # Создаем лист со статистикой
        ws_stats = wb.create_sheet("Статистика")
        self._create_stats_sheet(ws_stats, df)
        
        # Создаем лист с регионами
        ws_regions = wb.create_sheet("По регионам")
        self._create_regions_sheet(ws_regions, df)
        
        # Сохраняем файл
        wb.save(output_file)
        logger.info(f"XLSX файл сохранен: {output_file}")
        
        return output_file
        
    def export_from_db_to_xlsx(self, output_file: str = 'goszakup_full_database.xlsx'):
        """Экспорт из базы данных в XLSX"""
        logger.info(f"Экспорт из базы данных в {output_file}")
        
        conn = sqlite3.connect(self.db_file)
        
        # Получаем основные данные
        main_query = '''
            SELECT 
                participant_number as "Номер участника",
                name as "Наименование",
                bin as "БИН",
                iin as "ИИН",
                rnn as "РНН",
                supplier_id as "ID поставщика",
                detail_url as "Ссылка",
                is_parsed as "Данные получены",
                created_at as "Дата добавления",
                updated_at as "Дата обновления"
            FROM suppliers
            ORDER BY participant_number
        '''
        
        df_main = pd.read_sql_query(main_query, conn)
        
        # Получаем детальные данные
        details_query = '''
            SELECT 
                supplier_id,
                section as "Секция",
                field_name as "Поле",
                field_value as "Значение"
            FROM supplier_details
            ORDER BY supplier_id, section, field_name
        '''
        
        df_details = pd.read_sql_query(details_query, conn)
        conn.close()
        
        # Создаем workbook
        wb = Workbook()
        
        # Основной лист
        ws_main = wb.active
        ws_main.title = "Поставщики"
        
        # Записываем основные данные
        for r in dataframe_to_rows(df_main, index=False, header=True):
            ws_main.append(r)
            
        self._format_main_sheet(ws_main, len(df_main.columns))
        
        # Лист с детальными данными
        if not df_details.empty:
            ws_details = wb.create_sheet("Детальная информация")
            for r in dataframe_to_rows(df_details, index=False, header=True):
                ws_details.append(r)
            self._format_details_sheet(ws_details, len(df_details.columns))
        
        # Статистика
        ws_stats = wb.create_sheet("Статистика БД")
        self._create_db_stats_sheet(ws_stats, df_main, df_details)
        
        wb.save(output_file)
        logger.info(f"XLSX файл из БД сохранен: {output_file}")
        
        return output_file
        
    def _format_main_sheet(self, ws, num_columns):
        """Форматирование основного листа"""
        
        # Стили
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        data_font = Font(name='Arial', size=10)
        alternate_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        alignment_center = Alignment(horizontal='center', vertical='center')
        alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        
        # Форматируем заголовки
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = border
            
        # Форматируем данные и устанавливаем ширину колонок
        column_widths = [15, 40, 15, 15, 15, 12, 20, 25, 25, 12, 30, 30, 15, 12, 20, 20, 30, 15, 30, 40, 25, 15]
        
        for row in range(2, ws.max_row + 1):
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = border
                
                if col <= 6:  # Основные колонки - по центру
                    cell.alignment = alignment_center
                else:  # Остальные - по левому краю с переносом
                    cell.alignment = alignment_left
                
                # Чередующиеся цвета строк
                if row % 2 == 0:
                    cell.fill = alternate_fill
                    
        # Устанавливаем ширину колонок
        for i, width in enumerate(column_widths[:num_columns], 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
            
        # Замораживаем первую строку
        ws.freeze_panes = 'A2'
        
        # Добавляем автофильтр
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=num_columns).coordinate}"
        
    def _format_details_sheet(self, ws, num_columns):
        """Форматирование листа с деталями"""
        self._format_main_sheet(ws, num_columns)
        
    def _create_stats_sheet(self, ws, df):
        """Создание листа со статистикой"""
        ws.append(['СТАТИСТИКА ПАРСИНГА GOSZAKUP.GOV.KZ'])
        ws.append([])
        ws.append(['Дата создания отчета:', datetime.now().strftime('%d.%m.%Y %H:%M:%S')])
        ws.append([])
        
        # Общая статистика
        ws.append(['ОБЩАЯ ИНФОРМАЦИЯ'])
        ws.append(['Всего записей:', len(df)])
        ws.append(['Записей с БИН:', len(df[df['bin'].notna() & (df['bin'] != '')])])
        ws.append(['Записей с ИИН:', len(df[df['iin'].notna() & (df['iin'] != '')])])
        ws.append(['Записей с РНН:', len(df[df['rnn'].notna() & (df['rnn'] != '')])])
        ws.append([])
        
        # Статистика по регионам
        if 'detail_Регион' in df.columns:
            ws.append(['СТАТИСТИКА ПО РЕГИОНАМ'])
            region_stats = df['detail_Регион'].value_counts().head(10)
            for region, count in region_stats.items():
                ws.append([region, count])
        
        # Форматирование
        for row in range(1, ws.max_row + 1):
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                if row == 1:
                    cell.font = Font(size=16, bold=True)
                elif str(cell.value).isupper() and len(str(cell.value)) > 10:
                    cell.font = Font(size=12, bold=True, color='366092')
                else:
                    cell.font = Font(size=10)
                    
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
    def _create_regions_sheet(self, ws, df):
        """Создание листа с группировкой по регионам"""
        if 'detail_Регион' not in df.columns:
            ws.append(['Данные по регионам недоступны'])
            return
            
        region_data = df.groupby('detail_Регион').agg({
            'participant_number': 'count',
            'bin': lambda x: x.notna().sum(),
            'iin': lambda x: x.notna().sum(),
            'detail_E-Mail:': lambda x: x.notna().sum()
        }).reset_index()
        
        region_data.columns = ['Регион', 'Всего поставщиков', 'С БИН', 'С ИИН', 'С Email']
        
        # Записываем данные
        for r in dataframe_to_rows(region_data, index=False, header=True):
            ws.append(r)
            
        self._format_main_sheet(ws, len(region_data.columns))
        
    def _create_db_stats_sheet(self, ws, df_main, df_details):
        """Создание листа со статистикой БД"""
        ws.append(['СТАТИСТИКА БАЗЫ ДАННЫХ GOSZAKUP'])
        ws.append([])
        ws.append(['Дата создания отчета:', datetime.now().strftime('%d.%m.%Y %H:%M:%S')])
        ws.append([])
        
        # Статистика основной таблицы
        ws.append(['ОСНОВНАЯ ТАБЛИЦА ПОСТАВЩИКОВ'])
        ws.append(['Всего записей:', len(df_main)])
        ws.append(['Обработано детально:', len(df_main[df_main['Данные получены'] == 1])])
        ws.append(['Ожидают обработки:', len(df_main[df_main['Данные получены'] == 0])])
        ws.append([])
        
        # Статистика деталей
        if not df_details.empty:
            ws.append(['ДЕТАЛЬНАЯ ИНФОРМАЦИЯ'])
            ws.append(['Всего детальных записей:', len(df_details)])
            section_stats = df_details['Секция'].value_counts()
            for section, count in section_stats.items():
                ws.append([f'Секция "{section}":', count])
        
        # Форматирование
        for row in range(1, ws.max_row + 1):
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                if row == 1:
                    cell.font = Font(size=16, bold=True)
                elif str(cell.value).isupper() and len(str(cell.value)) > 10:
                    cell.font = Font(size=12, bold=True, color='366092')
                    
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

def main():
    """Главная функция экспорта"""
    exporter = XLSXExporter()
    
    print("=== ЭКСПОРТ В КРАСИВЫЙ XLSX ===")
    print()
    print("1. Экспорт демо данных (demo_suppliers.csv)")
    print("2. Экспорт из турбо базы данных")
    print("3. Экспорт из обычной базы данных")
    
    choice = input("Выберите опцию (1-3): ").strip()
    
    if choice == "1":
        if os.path.exists('demo_suppliers.csv'):
            output_file = exporter.export_demo_to_xlsx()
            print(f"✅ Демо данные экспортированы в: {output_file}")
        else:
            print("❌ Файл demo_suppliers.csv не найден!")
            
    elif choice == "2":
        if os.path.exists('turbo_goszakup.db'):
            output_file = exporter.export_from_db_to_xlsx()
            print(f"✅ Данные из турбо БД экспортированы в: {output_file}")
        else:
            print("❌ Турбо база данных не найдена!")
            
    elif choice == "3":
        if os.path.exists('goszakup_database.db'):
            exporter.db_file = 'goszakup_database.db'
            output_file = exporter.export_from_db_to_xlsx('goszakup_regular_export.xlsx')
            print(f"✅ Данные из обычной БД экспортированы в: {output_file}")
        else:
            print("❌ Обычная база данных не найдена!")
    else:
        print("❌ Неверный выбор!")

if __name__ == "__main__":
    import os
    main() 